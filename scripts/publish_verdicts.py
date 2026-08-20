#!/usr/bin/env python3
"""
Aristos verdicts publisher — deterministic formatter.

Reads the committed scout verdict JSONs and writes:
  1. dist/aristos_verdicts_<run_date>.xlsx  — ONE workbook, three named tabs
  2. dist/summary.json                      — small digest for the weekly notification
  3. (optional) pushes the workbook straight into a Google Sheet, if creds are present

THE MATH HAS ALREADY JUDGED. This script only reformats — it never recomputes,
reorders or reinterprets a verdict. Every cell traces to a field in the JSON.

Usage (inside the scout-verdicts GitHub Action, after the grader has committed):
    python scripts/publish_verdicts.py --reports reports/scout --out dist

Optional direct-to-Sheets push (no model tokens, no manual Copy-to):
    export GOOGLE_SERVICE_ACCOUNT_JSON="$(cat sa.json)"
    python scripts/publish_verdicts.py --reports reports/scout --out dist \
        --spreadsheet-id 1wDYdPDI_XDBTvx_ttmFskNbVpVDPonSJy_-auPY_Hb4 --append

Exit codes: 0 ok · 2 no sources found · 3 stale (with --max-age-days)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import pathlib
import sys
from collections import Counter

SOURCES = ["ft", "economist", "growth-scanner", "holdings"]

# internal lens id -> column header
LENS_COLUMNS = [
    ("conservative_plus_v1", "Defensive Income"),
    ("magic_formula_momentum_v1", "Value+Momentum"),
    ("growth_garp_v2", "GARP"),
    ("magic_formula_raw_v1", "Greenblatt RAW"),
    ("financials_v1", "Financials"),
]
LENS_ORDER = [lid for lid, _ in LENS_COLUMNS]
FRIENDLY = dict(LENS_COLUMNS)

NEWS_HEADERS = ["Run date", "Source", "Ticker", "Company", "Scouted on", "Best lens",
                *[h for _, h in LENS_COLUMNS], "Narrative", "F-Score"]
SCANNER_HEADERS = ["Run date", "Scan file", "Ticker", "Company", "Market", "6m return", "Best lens",
                   *[h for _, h in LENS_COLUMNS], "Scanner notes", "Narrative", "F-Score"]
HOLDINGS_HEADERS = ["Run date", "Ticker", "Company", "F-Score", "Best lens",
                    *[h for _, h in LENS_COLUMNS], "Narrative"]

MAX_CELL = 120


# ---------------------------------------------------------------- helpers

def load(reports: pathlib.Path, source: str) -> dict | None:
    p = reports / source / "latest.json"
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding="utf-8"))


def truncate(rendered: str) -> str:
    """Shorten a long rendered cell without ever touching the verdict or position."""
    if len(rendered) <= MAX_CELL:
        return rendered
    base = rendered.split(" [")[0]           # drop bracketed detail (flags, borderline notes)
    if len(base) <= MAX_CELL:
        return base + " […]"
    i = base.find("(")                        # then drop the parenthetical observed/threshold
    if i > 0:
        base = base[:i].rstrip() + " (…)"
    return base if len(base) <= MAX_CELL else base[: MAX_CELL - 1] + "…"


def short_reason(rendered: str) -> str:
    """Compress an exclusion into a short label for the narrative."""
    r = rendered
    if r.startswith("excluded — "):
        r = r[len("excluded — "):]
    r = r.split(" [")[0]
    if r.startswith("screen: "):
        return r[len("screen: "):].split(" (")[0] + " screen"
    if r.startswith("sector excluded"):
        return "sector excluded (Financial Services)"
    if r.startswith("asset kind") and "outside this strategy's scope" in r:
        return "asset-kind gate (" + r.split("'")[1] + ")"
    if "outside this strategy's scope" in r:
        return "sector out of scope"
    return r.split(" (")[0]


_NAME_STOPWORDS = {"the", "a", "an", "inc", "corp", "corporation", "group", "holding",
                   "holdings", "plc", "ltd", "limited", "company", "co", "sa", "ag", "nv",
                   "as", "de", "new", "common", "stock", "class"}


def _brand_token(company: str) -> str:
    """First distinctive word of the company name (skipping articles / corporate suffixes)."""
    import re
    head = re.split(r"[—\-–(]", company)[0]
    for t in re.findall(r"[A-Za-z][A-Za-z\.&]{2,}", head):
        if t.lower().strip(".") not in _NAME_STOPWORDS:
            return t
    return ""


def suspect_mismaps(payloads: dict) -> list[dict]:
    """Advisory only: the resolved `display` name doesn't contain the scouted company's
    distinctive word, which usually means the ticker resolved to the wrong security
    (e.g. Barrick Gold scouted, 'Gold.com, Inc.' resolved). Heuristic — it will miss
    subtle cases, so it is reported as an advisory and never written into a narrative."""
    out = []
    for source, data in payloads.items():
        for e in data["scouted"]:
            display, symbol, company = e.get("display", ""), e["symbol"], e.get("company", "")
            if not display or display == symbol or not company:
                continue                      # bare-ticker display carries no name to compare
            token = _brand_token(company)
            if token and token.lower() not in display.lower():
                out.append({"source": source, "symbol": symbol,
                            "company": company, "display": display})
    return out


def ranked_lenses(entry: dict) -> dict:
    """lens id -> (verdict, position, cohort_size) for every lens that actually ranked it."""
    return {lid: (c["verdict"], c["position"], c["cohort_size"])
            for lid, c in entry["lenses"].items() if c.get("status") == "ranked"}


def best_lens(ranked: dict):
    """Lowest position wins; ties broken by the canonical lens order."""
    if not ranked:
        return None, "no ranked lens — see cells"
    lid, (v, p, c) = min(ranked.items(), key=lambda kv: (kv[1][1], LENS_ORDER.index(kv[0])))
    return (lid, v, p, c), f"{v.upper()} — {lid}, #{p} of {c}"


def narrative(ranked: dict, cells: dict, extra: str | None = None) -> str:
    bits: list[str] = []
    diverging = sum(1 for r in cells.values() if "price diverging" in r)

    if ranked:
        (lid, v, p, c), _ = best_lens(ranked)
        s = f"Ranked by {len(ranked)} of 5 lenses; best is {FRIENDLY[lid]} at #{p} of {c} ({v.upper()})"
        others = [f"{FRIENDLY[l]} #{t[1]} of {t[2]} ({t[0].upper()})"
                  for l, t in ranked.items() if l != lid]
        if others:
            s += "; also " + ", ".join(others)
        bits.append(s + ".")

        verdicts = {t[0] for t in ranked.values()}
        if len(verdicts) > 1:
            bits.append("Lenses disagree (" + "/".join(sorted(x.upper() for x in verdicts)) + ").")

        # honesty flag: a BUY out of a tiny cohort is not a comparison
        if v == "buy" and c < 5:
            bits.append("Caution: this BUY is #1 of 1 — a cohort of one is not a comparison."
                        if c == 1 else
                        f"Caution: this BUY comes from a cohort of only {c} ranked names.")

        if len(ranked) < 5:
            cnt = Counter(short_reason(r) for l, r in cells.items() if l not in ranked)
            if cnt:
                bits.append("Elsewhere excluded: " +
                            ", ".join(k + (f" ×{n}" if n > 1 else "") for k, n in cnt.most_common(3)) + ".")
    else:
        cnt = Counter(short_reason(r) for r in cells.values())
        bits.append("No lens ranked it — excluded on " +
                    ", ".join(k + (f" ×{n}" if n > 1 else "") for k, n in cnt.most_common()) + ".")

    if diverging:
        bits.append(f"⚠ price-divergence flag on {diverging} lens screen(s).")
    if extra:
        bits.append(extra)
    return " ".join(bits)


def fscore_display(entry: dict) -> tuple[str, str | None]:
    """Returns (cell value, narrative fragment). Verbatim display string — never a bare number."""
    fs = entry.get("f_score")
    if not fs or not fs.get("display"):
        return "", None
    disp = fs["display"]
    unavailable = fs.get("unavailable", 0)
    if disp == "ABSTAIN":
        return disp, f"F-Score ABSTAIN ({unavailable} of 9 checks unavailable)."
    if unavailable:
        return disp, f"F-Score {disp} ({unavailable} check(s) unavailable)."
    return disp, f"F-Score {disp}."


# ---------------------------------------------------------------- tab builders

def build_news(payloads: dict) -> tuple[list[list], list[tuple]]:
    rows, stats = [NEWS_HEADERS], []
    for source, label in (("ft", "FT"), ("economist", "Economist")):
        data = payloads.get(source)
        if not data:
            continue
        for e in data["scouted"]:
            cells = {lid: e["lenses"][lid]["rendered"] for lid in LENS_ORDER}
            ranked = ranked_lenses(e)
            _, bl = best_lens(ranked)
            fs_cell, fs_note = fscore_display(e)
            rows.append([data["run_date"], label, e["symbol"], e["company"], e.get("date_added", ""), bl,
                         *[truncate(cells[lid]) for lid in LENS_ORDER],
                         narrative(ranked, cells, fs_note), fs_cell])
            stats.append((label, e["symbol"], bl, fs_cell))
    return rows, stats


def build_scanner(payloads: dict) -> tuple[list[list], list[tuple]]:
    rows, stats = [SCANNER_HEADERS], []
    data = payloads.get("growth-scanner")
    if not data:
        return rows, stats
    for e in data["scouted"]:
        cells = {lid: e["lenses"][lid]["rendered"] for lid in LENS_ORDER}
        ranked = ranked_lenses(e)
        _, bl = best_lens(ranked)
        notes = e.get("scanner_notes", "") or ""
        fs_cell, fs_note = fscore_display(e)

        if "no_prior_yr" in notes or "fundamentals_gt_300d_old" in notes:
            extra = (f"Scanner flags missing prior-year fundamentals ({notes}) — "
                     "shell/new-listing data artefact, not a signal.")
        elif notes:
            extra = f"Scanner notes: {notes}."
        else:
            extra = None
        if fs_note:
            extra = f"{extra} {fs_note}" if extra else fs_note

        nar = narrative(ranked, cells, extra)
        if ranked:
            nar = f"6m return {e.get('ret_6m')}. " + nar

        # scan_file lives on the payload; older runs carry it per-entry in date_added
        scan_file = data.get("scan_file") or e.get("date_added", "")
        rows.append([data["run_date"], scan_file, e["symbol"], e["company"], e.get("market", ""),
                     e.get("ret_6m", ""), bl,
                     *[truncate(cells[lid]) for lid in LENS_ORDER],
                     notes, nar, fs_cell])
        stats.append(("Scanner", e["symbol"], bl, fs_cell))
    return rows, stats


def build_holdings(payloads: dict) -> tuple[list[list], list[tuple]]:
    rows, stats = [HOLDINGS_HEADERS], []
    data = payloads.get("holdings")
    if not data:
        return rows, stats
    for e in data["scouted"]:
        cells = {lid: e["lenses"][lid]["rendered"] for lid in LENS_ORDER}
        ranked = ranked_lenses(e)
        _, bl = best_lens(ranked)
        fs_cell, fs_note = fscore_display(e)

        extra = None
        if e.get("holding_type") == "ETF":
            extra = ("ETF — excluded by the asset-kind gate in every stock lens; "
                     "correct behavior, not an error.")
        if fs_note:
            extra = f"{extra} {fs_note}" if extra else fs_note

        rows.append([data["run_date"], e["symbol"], e["company"], fs_cell, bl,
                     *[truncate(cells[lid]) for lid in LENS_ORDER],
                     narrative(ranked, cells, extra)])
        stats.append(("Holdings", e["symbol"], bl, fs_cell))
    return rows, stats


# ---------------------------------------------------------------- outputs

def write_xlsx(path: pathlib.Path, tabs: list[tuple[str, list[list]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in tabs:
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for i, header in enumerate(rows[0], start=1):
            width = {"Narrative": 70, "Company": 30, "Best lens": 34, "Scanner notes": 34}.get(header, 20)
            ws.column_dimensions[get_column_letter(i)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def push_to_sheets(spreadsheet_id: str, tabs: list[tuple[str, list[list]]], append: bool) -> None:
    """Write straight into the master spreadsheet. Requires GOOGLE_SERVICE_ACCOUNT_JSON."""
    import os
    from google.oauth2.service_account import Credentials      # google-auth
    from googleapiclient.discovery import build                 # google-api-python-client

    creds = Credentials.from_service_account_info(
        json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    api = build("sheets", "v4", credentials=creds).spreadsheets()
    existing = {s["properties"]["title"] for s in
                api.get(spreadsheetId=spreadsheet_id).execute()["sheets"]}

    from openpyxl.utils import get_column_letter

    for name, rows in tabs:
        header, data = rows[0], rows[1:]

        if name not in existing:                       # brand new tab: header + data at A1
            api.batchUpdate(spreadsheetId=spreadsheet_id,
                            body={"requests": [{"addSheet": {"properties": {"title": name}}}]}).execute()
            api.values().update(spreadsheetId=spreadsheet_id, range=f"'{name}'!A1",
                                valueInputOption="RAW", body={"values": rows}).execute()
            print(f"  created tab '{name}' with {len(data)} rows")
            continue

        if not append:                                 # overwrite in place
            api.values().update(spreadsheetId=spreadsheet_id, range=f"'{name}'!A1",
                                valueInputOption="RAW", body={"values": rows}).execute()
            print(f"  overwrote tab '{name}' with {len(data)} rows")
            continue

        # Appending to a tab that already holds prior weeks' rows. If the live header is
        # narrower than ours (e.g. the F-Score column was never added), extend it — write
        # ONLY the missing trailing cells so existing column names are never disturbed.
        live = api.values().get(spreadsheetId=spreadsheet_id,
                                range=f"'{name}'!1:1").execute().get("values", [[]])
        live_header = live[0] if live else []
        if len(live_header) < len(header):
            start = get_column_letter(len(live_header) + 1)
            end = get_column_letter(len(header))
            api.values().update(
                spreadsheetId=spreadsheet_id, range=f"'{name}'!{start}1:{end}1",
                valueInputOption="RAW",
                body={"values": [header[len(live_header):]]}).execute()
            print(f"  extended header of '{name}' with {header[len(live_header):]}")
        elif live_header and live_header[:len(header)] != header:
            print(f"  WARNING: '{name}' header differs from expected; appending anyway "
                  f"— verify column alignment", file=sys.stderr)

        api.values().append(spreadsheetId=spreadsheet_id, range=f"'{name}'!A1",
                            valueInputOption="RAW", insertDataOption="INSERT_ROWS",
                            body={"values": data}).execute()
        print(f"  appended {len(data)} rows to '{name}'")


def summarise(payloads: dict, all_stats: list[tuple], holdings_stats: list[tuple]) -> dict:
    """Small digest so the weekly notification needs no model reasoning over the raw rows."""
    def verdict_of(best_lens_str: str) -> str:
        return "NONE" if best_lens_str.startswith("no ranked lens") else best_lens_str.split(" — ")[0]

    per_source: dict[str, dict] = {}
    for label, ticker, bl, _fs in all_stats:
        d = per_source.setdefault(label, {"counts": Counter(), "buys": []})
        v = verdict_of(bl)
        d["counts"][v] += 1
        if v == "BUY":
            d["buys"].append(ticker)

    return {
        "run_dates": {s: p["run_date"] for s, p in payloads.items()},
        "per_source": {k: {"counts": dict(v["counts"]), "buys": v["buys"]} for k, v in per_source.items()},
        "holdings_f_scores": {t: fs for _l, t, _bl, fs in holdings_stats if fs},
        "holdings_abstains": [t for _l, t, _bl, fs in holdings_stats if fs == "ABSTAIN"],
        "sources_without_f_scores": [
            s for s, p in payloads.items()
            if not any(e.get("f_score") for e in p["scouted"])
        ],
        "skipped": {s: p.get("skipped", []) for s, p in payloads.items() if p.get("skipped")},
        "suspect_ticker_mismaps": suspect_mismaps(payloads),
        "scan_header": (payloads.get("growth-scanner") or {}).get("scan_header"),
        "notes": [
            "Holdings verdicts are RELATIVE ranks within the combined scout cohort — "
            "a SELL means 'ranks in the bottom of this cohort', never a sell instruction.",
            "News and scanner rank in different cohorts — a #5 in one is not comparable to a #5 in the other.",
            "Not investment advice.",
        ],
    }


# ---------------------------------------------------------------- main

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--reports", default="reports/scout", type=pathlib.Path)
    ap.add_argument("--out", default="dist", type=pathlib.Path)
    ap.add_argument("--max-age-days", type=int, default=None,
                    help="fail with exit 3 if the newest run_date is older than this")
    ap.add_argument("--spreadsheet-id", default=None, help="push tabs into this Google Sheet")
    ap.add_argument("--append", action="store_true",
                    help="append rows to existing tabs instead of overwriting")
    args = ap.parse_args()

    payloads = {s: d for s in SOURCES if (d := load(args.reports, s))}
    if not payloads:
        print(f"no source payloads under {args.reports}", file=sys.stderr)
        return 2
    for s in SOURCES:
        if s not in payloads:
            print(f"note: source '{s}' absent — publishing the others", file=sys.stderr)

    run_date = max(p["run_date"] for p in payloads.values())
    if args.max_age_days is not None:
        age = (dt.date.today() - dt.date.fromisoformat(run_date)).days
        if age > args.max_age_days:
            print(f"stale: newest run_date {run_date} is {age} days old", file=sys.stderr)
            return 3

    news_rows, news_stats = build_news(payloads)
    scan_rows, scan_stats = build_scanner(payloads)
    hold_rows, hold_stats = build_holdings(payloads)

    tabs = [("Aristos Verdicts", news_rows),
            ("Growth Scanner Verdicts", scan_rows),
            ("Holdings Verdicts", hold_rows)]

    xlsx = args.out / f"aristos_verdicts_{run_date}.xlsx"
    write_xlsx(xlsx, tabs)

    summary = summarise(payloads, news_stats + scan_stats + hold_stats, hold_stats)
    summary["workbook"] = xlsx.name
    summary["row_counts"] = {name: len(rows) - 1 for name, rows in tabs}
    (args.out / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    if args.spreadsheet_id:
        push_to_sheets(args.spreadsheet_id, tabs, args.append)
        print(f"pushed 3 tabs into spreadsheet {args.spreadsheet_id}")

    print(f"wrote {xlsx} ({summary['row_counts']}) and {args.out / 'summary.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
